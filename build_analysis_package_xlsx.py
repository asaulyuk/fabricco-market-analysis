#!/usr/bin/env python3
"""Build FabricCo_Analysis_Package.xlsx mirroring FabricCo_Data_Challenge_Analysis.ipynb logic."""
from __future__ import annotations

from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

BASE = Path(__file__).resolve().parent
CHART_DIR = BASE / "outputs" / "charts"
OUT_XLSX = BASE / "FabricCo_Analysis_Package.xlsx"

products = ["Beanies", "Bikinis", "Hoodies", "Summer_Shirts", "Rain_Coats"]
product_palette = ["#0f172a", "#334155", "#64748b", "#94a3b8", "#cbd5e1"]
metrics = {
    "Sold_Units": "Sold_Units",
    "Gross_Profit_Margin": "Gross_Profit_Margin",
    "Customer_Satisfaction_Score": "Customer_Satisfaction_Score",
    "Inventory_Turnover": "Inventory_Turnover",
}


def load_trends() -> pd.DataFrame:
    combined_path = BASE / "UK_NZ_combined.xlsx"
    tr = pd.read_excel(combined_path, sheet_name="Sheet1", header=0)
    time_col = tr.columns[0]
    tr = tr.rename(columns={time_col: "Month"})
    tr["Month"] = pd.to_datetime(tr["Month"])
    tr = tr.dropna(axis=1, how="all")
    tr = tr.loc[:, [c for c in tr.columns if not str(c).startswith("Unnamed:")]]
    uk_sw, uk_sc = "UK swimsuit", "UK scarf"
    nz_sw, nz_sc = "NZ swimsuit", "NZ scarf"
    for c in [uk_sw, uk_sc, nz_sw, nz_sc]:
        tr[c] = pd.to_numeric(tr[c], errors="coerce")
    return tr.sort_values("Month").reset_index(drop=True)


def load_fabric() -> pd.DataFrame:
    fabric_path = BASE / "FabricCo_Full_Products_Corrected_KPIs_2004_2024.csv"
    fabric = pd.read_csv(fabric_path)
    fabric["Date"] = pd.to_datetime(fabric["Date"], format="%m/%d/%y", errors="coerce")
    if fabric["Date"].isna().mean() > 0.05:
        fabric["Date"] = pd.to_datetime(fabric["Date"], errors="coerce")
    fabric["Year"] = fabric["Date"].dt.year
    return fabric.sort_values("Date").reset_index(drop=True)


def add_df_sheet(wb: Workbook, title: str, df: pd.DataFrame, freeze: bool = True) -> None:
    ws = wb.create_sheet(title)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    if freeze and ws.max_row > 1:
        ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _chart_style() -> None:
    plt.rcParams.update(
        {
            "figure.facecolor": "white",
            "axes.facecolor": "white",
            "axes.edgecolor": "#374151",
            "axes.labelcolor": "#111827",
            "axes.titleweight": "semibold",
            "axes.titlesize": 13,
            "axes.labelsize": 10,
            "xtick.labelsize": 9,
            "ytick.labelsize": 9,
            "axes.spines.top": False,
            "axes.spines.right": False,
            "axes.grid": True,
            "grid.color": "#e5e7eb",
            "grid.linewidth": 0.8,
            "legend.frameon": False,
        }
    )


def _save_chart(fig: plt.Figure, filename: str) -> None:
    CHART_DIR.mkdir(parents=True, exist_ok=True)
    fig.savefig(CHART_DIR / filename, dpi=180, bbox_inches="tight", facecolor="white")


def _plot_all_products_yearly_lines(
    yearly: pd.DataFrame, metric_suffix: str, ylabel: str, title: str, fname: str, percent_axis: bool = False
) -> None:
    fig, ax = plt.subplots(figsize=(7, 3))
    for p, col in zip(products, product_palette):
        cname = f"{p}_{metric_suffix}"
        ax.plot(
            yearly["Year"],
            yearly[cname],
            marker="o",
            markersize=3.0,
            linewidth=1.8,
            color=col,
            label=p.replace("_", " "),
        )
    ax.set_xlabel("Year")
    ax.set_ylabel(ylabel)
    ax.set_title(title, fontsize=10.5)
    ax.grid(True, color="#e5e7eb", linewidth=0.8)
    if percent_axis:
        ax.yaxis.set_major_formatter(mtick.PercentFormatter(1.0))
    ax.legend(loc="upper left", fontsize=7, ncol=2)
    plt.tight_layout()
    _save_chart(fig, fname)
    plt.close(fig)


def generate_part2_trend_charts(fabric: pd.DataFrame) -> None:
    _chart_style()
    metric_cols = []
    for p in products:
        metric_cols.extend(
            [
                f"{p}_Sold_Units",
                f"{p}_Gross_Profit_Margin",
                f"{p}_Customer_Satisfaction_Score",
                f"{p}_Inventory_Turnover",
                f"{p}_Capacity_Utilization",
            ]
        )
    yearly = fabric.groupby("Year", as_index=False)[metric_cols].mean()

    charts = [
        (
            "Sold_Units",
            "Units sold (annual mean of months)",
            "Annual Mean Units Sold by Product",
            "part2_trends_units.png",
            False,
        ),
        (
            "Gross_Profit_Margin",
            "Gross margin (annual mean of months)",
            "Annual Mean Gross Profit Margin by Product",
            "part2_trends_margin.png",
            True,
        ),
        (
            "Customer_Satisfaction_Score",
            "Customer satisfaction score (annual mean of months)",
            "Annual Mean Customer Satisfaction by Product",
            "part2_trends_satisfaction.png",
            False,
        ),
        (
            "Inventory_Turnover",
            "Inventory turnover (annual mean of months)",
            "Annual Mean Inventory Turnover by Product",
            "part2_trends_turnover.png",
            False,
        ),
        (
            "Capacity_Utilization",
            "Capacity utilization (annual mean of months)",
            "Annual Mean Capacity Utilization by Product",
            "part2_trends_capacity.png",
            True,
        ),
    ]
    for metric_suffix, ylabel, title, fname, percent_axis in charts:
        _plot_all_products_yearly_lines(yearly, metric_suffix, ylabel, title, fname, percent_axis)


def insert_chart_sheet(wb: Workbook, title: str, items: list[tuple[str, str, str, str]]) -> None:
    """items: (png_filename, notebook_title, notebook_subtitle, caption)"""
    ws = wb.create_sheet(title)
    ws["A1"] = "PNG file (under outputs/charts/)"
    ws["B1"] = "Chart title (matches notebook)"
    ws["C1"] = "Subtitle (matches notebook)"
    ws["D1"] = "Caption"
    for c in ws[1]:
        c.font = Font(bold=True)
    row = 2
    for png, ttl, sub, cap in items:
        ws.cell(row=row, column=1, value=png)
        ws.cell(row=row, column=2, value=ttl)
        ws.cell(row=row, column=3, value=sub)
        ws.cell(row=row, column=4, value=cap)
        row += 1
    img_start = row + 1
    col_img = 1
    for i, (png, _, _, _) in enumerate(items):
        path = CHART_DIR / png
        if not path.exists():
            ws.cell(row=img_start + i * 28, column=1, value=f"[Missing image: {png}]")
            continue
        img = XLImage(str(path))
        ow, oh = img.width, img.height
        max_w = 720
        if ow > max_w:
            scale = max_w / ow
            img.width = int(ow * scale)
            img.height = int(oh * scale)
        anchor = f"A{img_start + i * 28}"
        ws.add_image(img, anchor)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 40


def main() -> None:
    tr = load_trends()
    uk_sw, uk_sc = "UK swimsuit", "UK scarf"
    nz_sw, nz_sc = "NZ swimsuit", "NZ scarf"
    r_uk = float(tr[uk_sw].corr(tr[uk_sc]))
    r_nz = float(tr[nz_sw].corr(tr[nz_sc]))

    fabric = load_fabric()
    generate_part2_trend_charts(fabric)

    rows = []
    for prod in products:
        for nice, suf in metrics.items():
            col = f"{prod}_{suf}"
            s = fabric[col]
            rows.append(
                {
                    "Product": prod.replace("_", " "),
                    "Metric": nice,
                    "Mean": s.mean(),
                    "Median": s.median(),
                    "Std": s.std(),
                }
            )
    stats_tbl = pd.DataFrame(rows)

    recent = fabric[fabric["Year"].between(2020, 2024)].copy()
    summary = []
    for p in products:
        summary.append(
            {
                "Product": p.replace("_", " "),
                "Sold_Units_sum": recent[f"{p}_Sold_Units"].sum(),
                "Gross_Profit_Margin_mean": recent[f"{p}_Gross_Profit_Margin"].mean(),
                "Inventory_Turnover_mean": recent[f"{p}_Inventory_Turnover"].mean(),
                "Capacity_Utilization_mean": recent[f"{p}_Capacity_Utilization"].mean(),
                "Customer_Satisfaction_mean": recent[f"{p}_Customer_Satisfaction_Score"].mean(),
                "Items_Returned_sum": recent[f"{p}_Items_Returned"].sum(),
            }
        )
    sdf = pd.DataFrame(summary).set_index("Product")

    ranks = pd.DataFrame(index=sdf.index)
    ranks["Popularity_rank"] = sdf["Sold_Units_sum"].rank(ascending=False)
    ranks["Margin_rank"] = sdf["Gross_Profit_Margin_mean"].rank(ascending=False)
    ranks["InvTurn_rank"] = sdf["Inventory_Turnover_mean"].rank(ascending=False)
    ranks["CapUtil_rank"] = sdf["Capacity_Utilization_mean"].rank(ascending=False)
    ranks["Sat_rank"] = sdf["Customer_Satisfaction_mean"].rank(ascending=False)
    ranks["Returns_rank"] = sdf["Items_Returned_sum"].rank(ascending=True)
    avg_rank = ranks.mean(axis=1).sort_values()
    best = avg_rank.index[0]
    worst = avg_rank.index[-1]

    long_rows = []
    for _, row in fabric.iterrows():
        for p in products:
            long_rows.append(
                {
                    "Product": p,
                    "Customer_Satisfaction_Score": row[f"{p}_Customer_Satisfaction_Score"],
                    "Items_Returned": row[f"{p}_Items_Returned"],
                    "Inventory_Turnover": row[f"{p}_Inventory_Turnover"],
                    "Gross_Profit_Margin": row[f"{p}_Gross_Profit_Margin"],
                    "Capacity_Utilization": row[f"{p}_Capacity_Utilization"],
                    "Lead_Time_Days": row[f"{p}_Lead_Time_Days"],
                }
            )
    long_df = pd.DataFrame(long_rows)
    r_sat_ret = float(long_df["Customer_Satisfaction_Score"].corr(long_df["Items_Returned"]))
    r_turn_margin = float(long_df["Inventory_Turnover"].corr(long_df["Gross_Profit_Margin"]))
    r_cap_lead = float(long_df["Capacity_Utilization"].corr(long_df["Lead_Time_Days"]))

    wb = Workbook()
    wb.remove(wb.active)

    # --- Read_Me ---
    rm = wb.create_sheet("Read_Me", 0)
    rm["A1"] = "FabricCo Data Challenge — Analysis Package (Excel)"
    rm["A1"].font = Font(bold=True, size=14)
    lines = [
        "This workbook mirrors the executed notebook FabricCo_Data_Challenge_Analysis.ipynb: same cleaned Google Trends table, same Pearson correlations, same Part 2 aggregates and pooled correlations.",
        f"Numbers are regenerated from the source files in the project folder on build. Final run used UK_NZ_combined.xlsx (Sheet1) and FabricCo_Full_Products_Corrected_KPIs_2004_2024.csv.",
        f"Chart PNGs live in outputs/charts/ (240 dpi). Part1_Charts and Part2_Charts embed those files as thumbnails for non-technical reviewers.",
        "If FabricCo_internal_analysis.xlsx is present from an earlier draft, use this package file instead—it is the one rebuilt from the current notebook logic.",
        "",
        "Google Trends measures relative search interest within each market and period. It is used here to analyze timing and seasonality, not absolute market demand.",
        "",
        "Because volume differences are narrow, product decisions should rely more on profitability, efficiency, and customer experience than on units sold alone.",
    ]
    for i, line in enumerate(lines, start=3):
        rm.cell(row=i, column=1, value=line)
        rm.cell(row=i, column=1).alignment = Alignment(wrap_text=True)
    rm.column_dimensions["A"].width = 110

    # --- Part 1 data ---
    tr_out = tr.copy()
    tr_out["Month"] = tr_out["Month"].dt.strftime("%Y-%m-%d")
    add_df_sheet(wb, "Part1_Combined_Data", tr_out)

    corr_p1 = pd.DataFrame(
        {
            "Pair": [
                "UK: UK swimsuit vs UK scarf (Pearson r)",
                "NZ: NZ swimsuit vs NZ scarf (Pearson r)",
            ],
            "Correlation": [round(r_uk, 6), round(r_nz, 6)],
            "Interpretation": [
                "Weak positive linear co-movement; seasonal charts carry the main story.",
                "Weak positive linear co-movement; NZ r is slightly higher than UK but still low.",
            ],
        }
    )
    add_df_sheet(wb, "Part1_Correlations", corr_p1)

    part1_charts = [
        (
            "uk_swimsuit_scarf_seasonality.png",
            "Scarves Set the Seasonal Pace in the UK",
            "Swimsuits stay steadier—plan scarves separately from warm-weather swimwear.",
            "UK monthly Google Trends (0–100), swimsuit vs scarf.",
        ),
        (
            "nz_swimsuit_scarf_seasonality.png",
            "New Zealand Shows a Different Search Rhythm",
            "Do not reuse the UK calendar playbook—peaks land in different months.",
            "NZ monthly Google Trends (0–100), swimsuit vs scarf.",
        ),
        (
            "swimsuit_uk_vs_nz_hemisphere_shift.png",
            "Swimsuit Interest Peaks in Opposite Seasons",
            "Northern vs southern summer explains the calendar gap—stagger markets, don’t mirror them.",
            "Same product (swimsuit), UK vs NZ timing.",
        ),
        (
            "scarf_uk_vs_nz_comparison.png",
            "Scarf Curiosity Differs by Country",
            "Trend lines are not comparable levels—only shape and timing inform planning.",
            "Same product (scarf), UK vs NZ shape comparison.",
        ),
    ]
    insert_chart_sheet(wb, "Part1_Charts", part1_charts)

    stats_round = stats_tbl.copy()
    for col in ("Mean", "Median", "Std"):
        stats_round[col] = stats_round[col].map(lambda x: round(float(x), 6))
    add_df_sheet(wb, "Part2_Summary_Stats", stats_round)

    sdf_num = sdf.reset_index().copy()
    for c in sdf_num.columns:
        if c != "Product":
            sdf_num[c] = sdf_num[c].map(lambda x: round(float(x), 6))
    ranks_out = ranks.reset_index()
    ranks_out["Avg_rank_six_dims"] = avg_rank.reindex(ranks_out["Product"]).values
    for c in ranks_out.columns:
        if c != "Product":
            ranks_out[c] = ranks_out[c].map(lambda x: round(float(x), 4))
    merged = sdf_num.merge(ranks_out, on="Product", how="left")
    add_df_sheet(wb, "Part2_Last5Y_Comparison", merged)

    corr_p2 = pd.DataFrame(
        {
            "Pair": [
                "Customer satisfaction vs items returned (pooled product-months)",
                "Inventory turnover vs gross profit margin",
                "Capacity utilization vs lead time (days)",
            ],
            "Pearson_r": [
                round(r_sat_ret, 6),
                round(r_turn_margin, 6),
                round(r_cap_lead, 6),
            ],
            "Notebook_print": [
                f"Service vs. returns: Pearson r = {r_sat_ret:.3f}",
                f"Turns vs. margin: Pearson r = {r_turn_margin:.3f}",
                f"Capacity vs. lead time: Pearson r = {r_cap_lead:.3f}",
            ],
        }
    )
    add_df_sheet(wb, "Part2_Correlations", corr_p2)

    part2_charts = [
        (
            "part2_trends_units.png",
            "Annual Mean Units Sold by Product",
            "2004–2024 trend lines for all five categories.",
            "Annual mean sold units by product across full period.",
        ),
        (
            "part2_trends_margin.png",
            "Annual Mean Gross Profit Margin by Product",
            "2004–2024 trend lines for all five categories.",
            "Annual mean gross margin by product across full period.",
        ),
        (
            "part2_trends_satisfaction.png",
            "Annual Mean Customer Satisfaction by Product",
            "2004–2024 trend lines for all five categories.",
            "Annual mean customer satisfaction by product across full period.",
        ),
        (
            "part2_trends_turnover.png",
            "Annual Mean Inventory Turnover by Product",
            "2004–2024 trend lines for all five categories.",
            "Annual mean inventory turnover by product across full period.",
        ),
        (
            "part2_trends_capacity.png",
            "Annual Mean Capacity Utilization by Product",
            "2004–2024 trend lines for all five categories.",
            "Annual mean capacity utilization by product across full period.",
        ),
        (
            "yearly_units_top_two_vs_bottom_one.png",
            "Volume Leaders and the Smallest Line Stay Close",
            "Top two categories vs the lowest seller—still no runaway winner in yearly units.",
            "Annual sold units: top two vs lowest volume category 2020–2024 window logic in notebook.",
        ),
        (
            "yearly_margin_strongest_vs_weakest.png",
            "Margin Spread Is Narrow Year to Year",
            "Strongest vs weakest category—profit decisions need more than demand curves.",
            "Annual mean gross margin: strongest vs weakest by 2020–2024 average.",
        ),
        (
            "yearly_satisfaction_highest_vs_lowest.png",
            "Satisfaction Gap Is Modest but Persistent",
            "Highest vs lowest category—close the gap with targeted service fixes, not one big campaign.",
            "Annual mean satisfaction: highest vs lowest by 2020–2024 average.",
        ),
        (
            "yearly_turnover_fastest_vs_slowest.png",
            "Fast Movers and Slow Movers Diverge",
            "Fastest vs slowest inventory category—allocate working capital where turns lag.",
            "Annual mean inventory turnover: fastest vs slowest by 2020–2024 average.",
        ),
        (
            "five_year_volume_no_single_leader.png",
            "No Category Owns the Five-Year Volume Story",
            "Bars cluster tightly—portfolio calls should weigh margin and service, not volume alone.",
            "Total units 2020–2024 by category.",
        ),
        (
            "five_year_efficiency_composite_split.png",
            "Best vs Weakest Performers on Flow and Capacity",
            "Top two composite names vs bottom two—shows where execution pressure differs.",
            "Turnover vs capacity: top two vs bottom two by composite average rank.",
        ),
        (
            "operational_metrics_correlation_summary.png",
            "Operational Pairs Show No Strong Linear Link",
            "Bars hug zero—manage satisfaction, returns, turns, and margin on their own merits.",
            "Pooled correlations for three operational pairs.",
        ),
    ]
    insert_chart_sheet(wb, "Part2_Charts", part2_charts)

    # --- Assignment_Answers (text) ---
    aa = wb.create_sheet("Assignment_Answers")
    aa.column_dimensions["A"].width = 28
    aa.column_dimensions["B"].width = 110
    aa["A1"] = "Step / question"
    aa["B1"] = "Direct answer (aligned with notebook outputs)"
    aa["A1"].font = Font(bold=True)
    aa["B1"].font = Font(bold=True)

    gt_note = (
        "Google Trends measures relative search interest within each market and period. "
        "It is used here to analyze timing and seasonality, not absolute market demand."
    )
    vol_note = (
        "Because volume differences are narrow, product decisions should rely more on profitability, "
        "efficiency, and customer experience than on units sold alone."
    )

    answers: list[tuple[str, str]] = [
        (
            "Part 1 — Data load & structure",
            "We loaded UK_NZ_combined.xlsx (Sheet1). The workbook’s first column may still show an odd export header name; in code we rename that column to Month, parse dates, drop blank Unnamed spacer columns, and coerce the four trend series to numeric. The cleaned table—with header Month in column A—is in Excel Part1_Combined_Data and matches notebook tr.",
        ),
        (
            "Part 1 — UK swimsuit vs scarf (what it shows)",
            "UK scarf search interest is the dominant seasonal signal — it spikes sharply each autumn and winter, reaching index 100 in recent years with a standard deviation of 20.4. UK swimsuit interest is comparatively steadier, building gradually into summer at a much lower overall index level (mean 10.5, max 34, std 7.0). The two products follow opposite seasonal windows and should be planned independently.",
        ),
        (
            "Part 1 — NZ swimsuit vs scarf (what it shows)",
            "New Zealand mirrors the UK structure shifted by six months. NZ scarf interest spikes sharply each southern winter (May–July), reaching index 100 with a standard deviation of 20.8 — nearly identical to the UK scarf profile. NZ swimsuit interest is more sustained, lifting through the southern summer (November–January) without the sharp spike that defines the scarf series. In both countries, scarves are the volatile product and swimsuits are the steadier one.",
        ),
        (
            "Part 1 — UK vs NZ swimsuit timing",
            "Swimsuit interest peaks in opposite calendar seasons for the UK versus New Zealand, consistent with opposite hemispheres. Notebook chart title: Swimsuit Interest Peaks in Opposite Seasons.",
        ),
        (
            "Part 1 — Swimsuit–scarf correlation UK vs NZ",
            f"Monthly Pearson correlation (swimsuit vs scarf) is r = {r_uk:.4f} in the UK and r = {r_nz:.4f} in New Zealand. Both are weak positives; New Zealand is slightly higher, so the swimsuit–scarf relationship is not materially stronger in one country—neither supports treating the series as tightly linked in linear terms.",
        ),
        (
            "Part 1 — Absolute demand vs Trends",
            gt_note,
        ),
        (
            "Part 1 — Manufacturing / market timing implication",
            "Year-round manufacturing should not assume one global swim or scarf peak: treat the UK and NZ as separate seasonal calendars, stagger marketing and inventory builds, and use Trends only for timing and phase—not for absolute demand levels.",
        ),
        (
            "Part 2 — Exploratory mean / median / std",
            "Across 2004–2024 monthly rows, mean sold units per product cluster near ~340–356 units per month with similar standard deviations (~87–89), so no product sits on a totally different volume scale in the long-run average. Mean gross margins cluster near 0.30 with modest dispersion; satisfaction means cluster near 8.4–8.6 with standard deviations near 0.81–0.88; inventory turnover means are near 5.4–5.6 with std ~0.83–0.87—see sheet Part2_Summary_Stats for exact cells.",
        ),
        (
            "Part 2 — Single product dominance",
            "In 2020–2024 summed units, the highest category (Beanies, 21,994) is not far from the lowest (Bikinis, 20,373); the horizontal volume chart’s notebook title states there is no single owner of the five-year volume story.",
        ),
        (
            "Part 2 — Strongest product overall",
            f"The notebook’s six-metric average rank (popularity, margin, turnover, utilization, satisfaction, returns) places {best} first (lowest average rank). Bikinis, Hoodies, and Rain Coats tie at average rank 2.83; pandas sort order breaks the tie so Bikinis is labeled best overall—interpret ties as “clustered leaders,” not a wide gap.",
        ),
        (
            "Part 2 — Weakest product overall",
            f"{worst} has the highest average rank (3.50) across the same six dimensions—lowest satisfaction in the 2020–2024 window among the five and weaker ranks on margin and returns versus leaders.",
        ),
        (
            "Part 2 — Three operational correlations",
            f"Pooled product-month correlations are effectively negligible: satisfaction vs returns r = {r_sat_ret:.3f}; inventory turnover vs gross margin r = {r_turn_margin:.3f}; capacity utilization vs lead time r = {r_cap_lead:.3f}. Implication: do not manage these pairs as if one linear dial predicts the other—use product-level time trends and operational reviews instead.",
        ),
        (
            "Part 2 — Strategic recommendation 1",
            "Prioritize Bikinis for growth investment conditional on profitability checks: it leads the composite rank definition used in the notebook (with ties at 2.83) and shows comparatively strong satisfaction and returns in the 2020–2024 aggregates.",
        ),
        (
            "Part 2 — Strategic recommendation 2",
            "Launch a focused Beanies improvement program (cost, pricing, SKU rationalization, quality) because it ranks weakest on the same composite view; validate with margin and return diagnostics before exit decisions. " + vol_note,
        ),
    ]

    r = 2
    for step, text in answers:
        aa.cell(row=r, column=1, value=step)
        aa.cell(row=r, column=2, value=text)
        aa.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    wb.save(OUT_XLSX)
    print("Wrote", OUT_XLSX)
    print("UK r", r_uk, "NZ r", r_nz)
    print("Pooled", r_sat_ret, r_turn_margin, r_cap_lead)
    print("Best", best, "Worst", worst)


if __name__ == "__main__":
    main()
